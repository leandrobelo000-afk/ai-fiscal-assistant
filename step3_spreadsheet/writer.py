"""
step3_spreadsheet/writer.py
---------------------------
Grava os dados extraídos pela Etapa 2 em uma planilha Excel acumulativa.

Estrutura da planilha:
  - Aba "Resumo"  → uma linha por nota fiscal processada
  - Aba "Itens"   → todos os itens de todas as notas, com referência ao número da NF

A planilha é criada na primeira execução e acumulada nas seguintes —
novas notas são sempre adicionadas sem sobrescrever as anteriores.
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def verificar_dependencias():
    import importlib, subprocess
    for lib in ["pandas", "openpyxl"]:
        try:
            importlib.import_module(lib)
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", lib], check=True)


# ─────────────────────────────────────────────
# Configuração
# ─────────────────────────────────────────────

def selecionar_caminho_planilha() -> str:
    """Abre seletor para escolher onde salvar ou abrir a planilha."""
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()

    # Se já existe uma planilha, pergunta se quer abrir ela ou escolher outra
    caminho = filedialog.asksaveasfilename(
        title="Salvar planilha em...",
        defaultextension=".xlsx",
        initialfile="notas_fiscais.xlsx",
        filetypes=[("Excel", "*.xlsx")]
    )
    return caminho if caminho else None

CAMINHO_PLANILHA = selecionar_caminho_planilha()
if not CAMINHO_PLANILHA:
    print("Nenhuma pasta selecionada. Encerrando.")
    sys.exit(1)

COLUNAS_RESUMO = [
    "Data Processamento",
    "Número NF",
    "Data Emissão",
    "CNPJ Emitente",
    "Nome Emitente",
    "Valor Total (R$)",
    "Impostos (R$)",
    "Qtd Itens",
]

COLUNAS_ITENS = [
    "Número NF",
    "Data Emissão",
    "Nome Emitente",
    "Descrição Item",
    "Valor Item (R$)",
]


# ─────────────────────────────────────────────
# Criação inicial da planilha
# ─────────────────────────────────────────────

COR_CABECALHO   = "2D6A9F"   # azul
COR_CABECALHO_2 = "1A5276"   # azul escuro para aba de itens
COR_FONTE       = "FFFFFF"   # branco


def _estilizar_cabecalho(ws, colunas: list, cor: str):
    """Aplica estilo ao cabeçalho de uma aba."""
    for col_idx, nome in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx, value=nome)
        celula.font      = Font(bold=True, color=COR_FONTE, name="Arial", size=11)
        celula.fill      = PatternFill("solid", start_color=cor)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(nome) + 4, 16)
    ws.row_dimensions[1].height = 28

# ─────────────────────────────────────────────
# Verificação de duplicata
# ─────────────────────────────────────────────

def ja_existe(numero_nf: str) -> bool:
    """Verifica se a nota já foi registrada na planilha."""
    if not os.path.exists(CAMINHO_PLANILHA):
        return False
    try:
        df = pd.read_excel(CAMINHO_PLANILHA, sheet_name="Resumo")
        if "Número NF" in df.columns:
            return str(numero_nf) in df["Número NF"].astype(str).values
    except Exception:
        pass
    return False


# ─────────────────────────────────────────────
# Gravação dos dados
# ─────────────────────────────────────────────

def gravar_nota(nota) -> bool:
    """
    Recebe um objeto NotaFiscalProcessada (Etapa 2) e grava na planilha.
    Retorna True se gravou, False se a nota já existia.
    """
    # Verifica duplicata
    if nota.numero and ja_existe(nota.numero):
        print(f"  ⚠  NF {nota.numero} já registrada — ignorando.")
        return False

    # Cria planilha se não existir
    if not os.path.exists(CAMINHO_PLANILHA):
        criar_planilha()

    wb = load_workbook(CAMINHO_PLANILHA)
    ws_resumo = wb["Resumo"]
    ws_itens  = wb["Itens"]

    # Linha de resumo
    linha_resumo = [
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        nota.numero,
        nota.data_emissao,
        nota.cnpj_emitente,
        nota.nome_emitente,
        nota.valor_total,
        nota.impostos if nota.impostos is not None else 0.0,
        len(nota.itens),
    ]
    ws_resumo.append(linha_resumo)

    # Zebra stripes na aba Resumo
    ultima_linha = ws_resumo.max_row
    if ultima_linha % 2 == 0:
        cor_linha = "EBF5FB"
        for col in range(1, len(COLUNAS_RESUMO) + 1):
            ws_resumo.cell(ultima_linha, col).fill = PatternFill("solid", start_color=cor_linha)

    # Linhas de itens
    for item in nota.itens:
        ws_itens.append([
            nota.numero,
            nota.data_emissao,
            nota.nome_emitente,
            item.get("descricao"),
            item.get("valor"),
        ])

    # Fórmula de total na aba Resumo (coluna F = Valor Total)
    # Atualiza célula de total geral logo abaixo dos dados
    total_row = ws_resumo.max_row + 2
    ws_resumo.cell(total_row, 5, "TOTAL GERAL").font = Font(bold=True, name="Arial")
    ws_resumo.cell(total_row, 6, f"=SUM(F2:F{ws_resumo.max_row - 1})").font = Font(bold=True, name="Arial")

    wb.save(CAMINHO_PLANILHA)
    print(f"  ✅ NF {nota.numero or 'sem número'} gravada com sucesso.")
    return True

"""Cria a planilha com as duas abas e cabeçalhos formatados."""
def criar_planilha():
    wb = Workbook()

    # Aba Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    _estilizar_cabecalho(ws_resumo, COLUNAS_RESUMO, COR_CABECALHO)

    # Aba Itens
    ws_itens = wb.create_sheet("Itens")
    _estilizar_cabecalho(ws_itens, COLUNAS_ITENS, COR_CABECALHO_2)

    wb.save(CAMINHO_PLANILHA)
    print(f"  📊 Planilha criada em: {CAMINHO_PLANILHA}")


# ─────────────────────────────────────────────
# Execução direta
# ─────────────────────────────────────────────

if __name__ == "__main__":
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from step1_ocr.extractor import verificar_tesseract
    from step2_llm.processor import processar_nota_fiscal
    import tkinter as tk
    from tkinter import filedialog

    verificar_tesseract()

    # Permite selecionar múltiplos arquivos de uma vez
    root = tk.Tk()
    root.withdraw()
    arquivos = filedialog.askopenfilenames(
        title="Selecione uma ou mais notas fiscais",
        filetypes=[("Arquivos suportados", "*.pdf *.jpg *.jpeg *.png *.tiff *.bmp")]
    )

    if not arquivos:
        print("Nenhum arquivo selecionado.")
        sys.exit(1)

    print(f"\n📂 {len(arquivos)} arquivo(s) selecionado(s)")
    print("=" * 50)

    gravadas = 0
    ignoradas = 0

    for caminho in arquivos:
        print(f"\n📄 {os.path.basename(caminho)}")
        print("-" * 40)

        try:
            nota = processar_nota_fiscal(caminho)

            print(f"  Número : {nota.numero}")
            print(f"  Data   : {nota.data_emissao}")
            print(f"  CNPJ   : {nota.cnpj_emitente}")
            print(f"  Valor  : R$ {nota.valor_total}")
            print(f"  Rota   : {nota.rota} | Confiança: {nota.confianca}")

            resultado = gravar_nota(nota)
            if resultado:
                gravadas += 1
            else:
                ignoradas += 1

        except Exception as e:
            print(f"  ❌ Erro ao processar: {e}")

    print("\n" + "=" * 50)
    print(f"✅ Gravadas : {gravadas}")
    print(f"⚠  Ignoradas: {ignoradas} (duplicatas)")
    print(f"📊 Planilha : {CAMINHO_PLANILHA}")

